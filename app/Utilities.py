import PyPDF2
import docx
import json
import pandas as pd
import re
import openpyxl
import os
import shutil
from typing import Dict, Any
# logger_setup.py
import logging

logger = logging.getLogger("myapp")
logging.basicConfig(level=logging.INFO)




def extract_text_from_pdf(uploaded_pdf):
    """Extracts text from a PDF file."""
    pdf_reader = PyPDF2.PdfReader(uploaded_pdf)
    text = "\n".join([page.extract_text() for page in pdf_reader.pages if page.extract_text()])
    return text

def extract_text_from_word(uploaded_word):
    """Extracts text from a Word document."""
    doc = docx.Document(uploaded_word)
    text = "\n".join([para.text for para in doc.paragraphs if para.text])
    return text

def generate_prompt(pdf_text, word_text):
    return f"""
Return a single valid JSON object only. Do not include markdown, explanations, thinking steps, or code blocks.

Extract the following fields from the applicant's CV and job application text:

- "Full-name"
- "Date-of-birth"
- "Gender"
- "Nationality"
- "Country-Contact"
- "E-Mail"
- "Phone-number"
- "Holds-Master-Degree"
- "Year-of-graduation-Master"
- "Languages"
- "Skills and competences"
- "Research Experience"
- "Holds-Doctoral-Degree"
- "Fits mobility rules?"
- "English Proficiency?"
- "Visa required?"

**Rules**:
- If a field is missing or unclear, write `"Unknown"`.
- The values for the following must be `"Yes"` or `"No"`:
  - Holds-Master-Degree
  - Holds-Doctoral-Degree
  - Fits mobility rules?
  - English Proficiency?
  - Visa required?
- Set `"Visa required?"` to `"No"` if the nationality is from any EU country (e.g., France, Germany, etc.); otherwise, `"Yes"`.
- Set `"English Proficiency?"` to `"Yes"` only if C1/C2, IELTS ≥ 6.5, or TOEFL ≥ 90 is mentioned.
- Do **not** include any lists, nested objects, or numbered keys. All values must be strings.

Return ONLY a JSON object in this structure:

{{
  "Full-name": "...",
  "Date-of-birth": "...",
  "Gender": "...",
  "Nationality": "...",
  "Country-Contact": "...",
  "E-Mail": "...",
  "Phone-number": "...",
  "Holds-Master-Degree": "...",
  "Year-of-graduation-Master": "...",
  "Languages": "...",
  "Skills and competences": "...",
  "Research Experience": "...",
  "Holds-Doctoral-Degree": "...",
  "Fits mobility rules?": "...",
  "English Proficiency?": "...",
  "Visa required?": "..."
}}

Begin the response with `curly braces` and return nothing else.

---

CV TEXT:
{pdf_text}

JOB APPLICATION TEXT:
{word_text}
"""


def fix_trailing_commas(json_text):
    """Removes trailing commas before } or ] to fix JSON format."""
    return re.sub(r',\s*([\]}])', r'\1', json_text)

def fix_unclosed_brackets(json_text):
    """Ensures all opening { or [ have matching closing } or ]."""
    open_curly, close_curly = json_text.count('{'), json_text.count('}')
    open_square, close_square = json_text.count('['), json_text.count(']')

    json_text += '}' * (open_curly - close_curly)
    json_text += ']' * (open_square - close_square)
    
    return json_text

def get_json(response_text):
    """Extracts and fixes JSON before parsing."""
    # First try to find JSON inside code blocks
    json_match = re.search(r'```(?:json)?\s*(\{(?:[^{}]*|\{.*?\})*\})\s*```', response_text, re.DOTALL)

    if not json_match:
        # Fallback to finding raw JSON
        json_match = re.search(r'\{(?:[^{}]*|\{.*?\})*\}', response_text, re.DOTALL)

    if json_match:
        json_text = json_match.group(1) if '```' in json_match.group(0) else json_match.group(0)

        #Apply fixes before parsing
        json_text = fix_trailing_commas(json_text)
        json_text = fix_unclosed_brackets(json_text)

        try:
            return json.loads(json_text)
        except json.JSONDecodeError as e:
            print(f"Error decoding JSON: {e}")
            print(f"Problematic snippet: {json_text[max(0, e.pos - 40): min(len(json_text), e.pos + 40)]}")
            return None
    else:
        print("No valid JSON block found in the response.")
        return None


def flatten_json(nested_json, parent_key='', sep='_'):
    """
    Recursively flattens a nested JSON structure.

    Args:
        nested_json (dict or list): The JSON data.
        parent_key (str): Key prefix for nested fields.
        sep (str): Separator for flattened keys.

    Returns:
        dict: Flattened dictionary.
    """
    flattened_dict = {}

    def recurse(data, prefix=''):
        if isinstance(data, dict):
            for key, value in data.items():
                new_key = f"{prefix}{sep}{key}" if prefix else key
                recurse(value, new_key)
        elif isinstance(data, list):
            for i, item in enumerate(data):
                new_key = f"{prefix}{sep}{i}" if prefix else str(i)
                recurse(item, new_key)
        else:
            flattened_dict[prefix] = data

    recurse(nested_json)
    return flattened_dict

def normalize_text(text):
    if not text:
        return ""
    if not isinstance(text, str):
        text = str(text)
    return re.sub(r'\s+', ' ', text.lower().strip().replace("\n", " "))

def extract_combined_headers(ws, header_rows=[4, 5, 6]):
    headers = []
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        combined = []
        for row in header_rows:
            val = ws.cell(row=row, column=col).value
            if val:
                combined.append(str(val).strip())
        headers.append(" ".join(combined).strip() if combined else None)
    return headers

def check_english_proficiency_from_text(text):
    if not text:
        return "Filled manually"

    text = str(text).lower()

    if "c1" in text or "c2" in text:
        return "Yes"

    patterns = {
        'ielts': 6.5,
        'toefl': 90,
        'pte': 61,
        'duolingo': 110,
        'cambridge': 180,
        'cae': 180
    }

    for exam, threshold in patterns.items():
        match = re.search(rf"{exam}[^0-9]*([\d\.]+)", text)
        if match:
            try:
                score = float(match.group(1))
                if score >= threshold:
                    return "Yes"
            except:
                pass

    return "Filled manually"

def detect_master_degree(flat_json):
    for raw_key, raw_value in flat_json.items():
        key = normalize_text(raw_key)
        value = str(raw_value).strip().lower()

        # First check: if value is "yes" or "true" (any casing)
        if value in ["yes", "true", "1"]:
            return "Yes"

        # Second check: if key has "degree" and value has "master"
        if "degree" in key and "master" in value:
            return "Yes"

    # If neither found
    return "Filled manually"

def detect_doctoral_degree(flat_json):
    for key, value in flat_json.items():
        key_norm = normalize_text(key)
        value_str = str(value).strip().lower()

        if "phd" in key_norm or "phd" in value_str or "doctor" in value_str:
            if value_str in ["yes", "true", "1"]:
                return "Yes"
    return "No"

# List of European countries
european_countries = [
    "austria", "belgium", "bulgaria", "croatia", "cyprus", "czech republic",
    "denmark", "estonia", "finland", "france", "germany", "greece",
    "hungary", "ireland", "italy", "latvia", "lithuania", "luxembourg",
    "malta", "netherlands", "poland", "portugal", "romania", "slovakia",
    "slovenia", "spain", "sweden", "norway", "switzerland", "iceland",
    "liechtenstein"
]

def detect_visa_required(flat_json):
    visa_key = "visa required"
    nationality_key = "nationality"

    if visa_key in flat_json:
        value = str(flat_json[visa_key]).strip().lower()
        if value in ["yes", "true", "1"]:
            return "Yes"
        elif value in ["no", "false", "0"]:
            return "No"
        else:
            return "Filled manually"
    
    nationality = flat_json.get(nationality_key, "").lower()
    if any(country in nationality for country in european_countries):
        return "No"
    else:
        return "Yes"

def split_full_name(full_name):
    parts = full_name.strip().split()
    if not parts:
        return "", ""
    return parts[0], " ".join(parts[1:]) if len(parts) > 1 else ""

# --- Field mapping ---

field_map = {
    "first name": ["first name", "full-name", "full name", "name"],
    "last name": ["last name", "surname"],
    "gender": ["gender", "sex"],
    "date of birth": ["date-of-birth", "dob", "birthdate"],
    "nationality": ["nationality", "country-contact", "country", "country of origin"],
    "contact details (e.g. email)": ["email", "e-mail", "contact_email"],
    "phone number": ["phone_number", "phone", "phone-number", "mobile", "contact number"],
    "fits mobility rules?": ["mobility rule compliance", "mobility-rule-compliance", "fits mobility rules", "mobility rules"],
    "holds doctoral degree?": ["holds doctoral degree", "holds-doctoral-degree", "has phd", "phd", "doctorate"],
    "visa required?": ["visa required", "visa_required", "requires visa", "needs visa"],
    "english proficiency?": ["english proficiency", "language level", "english level", "spoken english"],
    "skills and competencies (max 5 points)": ["skills", "technical skills", "competencies", "frameworks", "tools", "programming-languages"],
    "holds/will hold a master degree or equivalent before october 2025?": ["master degree", "master-degree", "degree title", "holds-master-degree"]
}

# --- Excel Injection Function ---

def inject_standardized_json_to_excel(json_data: Dict[str, Any], template_path: str, output_path: str) -> None:
    """
    Injects standardized JSON data into an Excel file, appending to existing data.
    
    This function takes JSON data extracted from CVs and application forms, standardizes it,
    and appends it to an existing Excel file. If the output file doesn't exist, it creates
    a new one from the template.
    
    Args:
        json_data (Dict[str, Any]): The JSON data to inject, containing applicant information
        template_path (str): Path to the Excel template file
        output_path (str): Path where the Excel file should be saved/appended to
        
    Raises:
        FileNotFoundError: If the template file doesn't exist
        ValueError: If the JSON data is invalid or missing required fields
        Exception: For any other errors during Excel operations
    """
    # If output file doesn't exist, copy from template
    if not os.path.exists(output_path):
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template file not found at {template_path}")
        shutil.copy(template_path, output_path)
        logger.info(f"Created new output file from template at: {output_path}")

    try:
        # Load the workbook
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # Get headers
        headers = extract_combined_headers(ws)
        normalized_headers = [normalize_text(h) if h else "" for h in headers]

        # Flatten the JSON data
        flat_json = {}
        for k, v in json_data.items():
            if isinstance(v, dict):
                for subk, subv in v.items():
                    flat_json[f"{normalize_text(k)}_{normalize_text(subk)}"] = subv
            else:
                flat_json[normalize_text(k)] = v

        # Extract name information
        full_name = json_data.get("Full-name", "") or json_data.get("Name", "")
        if not full_name:
            raise ValueError("Missing required field: Full-name")
        first_name, last_name = split_full_name(full_name)

        # Extract language information
        language_info = json_data.get("Languages", "")
        if isinstance(language_info, dict):
            language_info = " ".join(language_info.values())

        # Find the next empty row
        row = 7  # Start after headers
        while ws.cell(row=row, column=1).value:
            row += 1

        # Write data to each column
        for col_idx, header in enumerate(headers, 1):
            if not header:
                ws.cell(row=row, column=col_idx, value="Filled manually")
                continue

            norm_header = normalize_text(header)
            value = "Filled manually"

            # Map fields based on header
            for expected, aliases in field_map.items():
                expected_norm = normalize_text(expected)
                if expected_norm in norm_header:
                    for alias in aliases:
                        alias = normalize_text(alias)
                        for flat_key, flat_val in flat_json.items():
                            if alias in flat_key:
                                value = flat_val
                                break
                        if value != "Filled manually":
                            break
                    break

            # Special handling for specific fields
            if "first name" in norm_header:
                value = first_name
            elif "last name" in norm_header:
                value = last_name
            elif "english proficiency" in norm_header:
                value = check_english_proficiency_from_text(language_info)
            elif "holds/will hold a master degree" in norm_header:
                value = detect_master_degree(flat_json)
            elif "holds doctoral degree?" in norm_header:
                value = detect_doctoral_degree(flat_json)
            elif "visa required?" in norm_header:
                value = detect_visa_required(flat_json)

            # Handle list values
            if isinstance(value, list):
                value = "; ".join(map(str, value))

            # Write the value to the cell
            ws.cell(row=row, column=col_idx, value=value)

        # Save the workbook
        wb.save(output_path)
        logger.info(f"Data successfully appended to row {row} in {output_path}")

    except Exception as e:
        logger.error(f"Error injecting data to Excel: {str(e)}")
        raise
